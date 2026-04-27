from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from ..adapters.bitrix_links import task_url
from ..config import Settings
from ..domain.candidates import CandidatePool
from ..domain.spec import Spec


_CATEGORY_RU = {
    "cameras": "Камеры",
    "conference": "Аудио / ВКС",
    "audio": "Аудио / ВКС",
    "display": "Отображение / Панели",
    "signal_transport": "Коммутация / Кабельная система",
    "processing": "Обработка / Управление",
    "misc": "Прочее",
}


def _task_link(tid: int, pool: Optional[CandidatePool], s: Settings) -> str:
    if pool:
        m = {t.task_id: t for t in pool.tasks}
        if tid in m and m[tid].url:
            return m[tid].url
    return task_url(tid, s)


def _md_cell(v: object) -> str:
    text = "" if v is None else str(v)
    # Escape markdown table separators and preserve multiline values.
    return text.replace("\\", "\\\\").replace("|", "\\|").replace("\n", "<br>")


def render_markdown(spec: Spec, *, pool: Optional[CandidatePool] = None, settings: Optional[Settings] = None) -> str:
    s = settings or Settings()

    lines: list[str] = []
    lines.append(f"# {spec.project_title}")

    if spec.project_summary:
        lines.append("\n## 1) Краткий вывод")
        lines.append(spec.project_summary)

    lines.append("\n## 2) Спецификация (СкСп) — шаблон")
    lines.append("| № | Производитель | Артикул | Описание | Цена, Р | Кол-во, шт | Сумма, ₽ | Комментарий |")
    lines.append("|---:|---|---|---|---:|---:|---:|---|")

    # group by category
    by_cat: dict[str, list] = {}
    for it in spec.items:
        by_cat.setdefault(it.category or "misc", []).append(it)

    n = 0
    for cat, items in by_cat.items():
        cat_ru = _CATEGORY_RU.get(cat, cat)
        lines.append(f"|  | **{cat_ru}** |  |  |  |  |  |  |")
        for it in items:
            n += 1
            price = it.unit_price.amount if it.unit_price else None
            total = (price * it.qty) if price is not None else None
            ev_ids = (it.evidence.bitrix_task_ids or [])[:2]
            ev = " ; ".join(_task_link(tid, pool, s) for tid in ev_ids) if ev_ids else ""
            comment = ev
            lines.append(
                f"| {n} | {_md_cell(it.manufacturer or '')} | {_md_cell(it.sku or '')} | {_md_cell(it.description or it.name)} | "
                f"{_md_cell(str(price) if price is not None else '')} | {_md_cell(it.qty)} | {_md_cell(str(total) if total is not None else '')} | {_md_cell(comment)} |"
            )

    lines.append("\n## 3) Почему выбран такой состав")
    lines.extend([f"- {x}" for x in (spec.why_composition or ["—"])])

    lines.append("\n## 4) Почему такие количества и цены")
    lines.extend([f"- {x}" for x in (spec.why_qty_and_price or ["—"])])

    lines.append("\n## 5) Прецеденты Bitrix (похожие задачи)")
    if spec.used_bitrix_task_ids:
        for tid in spec.used_bitrix_task_ids[:12]:
            lines.append(f"- {tid} — {_task_link(tid, pool, s)}")
    else:
        lines.append("- —")

    lines.append("\n## 6) Вопросы менеджера к заказчику")
    lines.extend([f"- {q}" for q in (spec.manager_questions or ["—"])])

    lines.append("\n## 7) Допущения и риски")
    if spec.assumptions:
        lines.append("**Допущения:**")
        lines.extend([f"- {a}" for a in spec.assumptions])
    if spec.risks:
        lines.append("\n**Риски:**")
        lines.extend([f"- {r}" for r in spec.risks])
    if not spec.assumptions and not spec.risks:
        lines.append("- —")

    return "\n".join(lines) + "\n"


def export_xlsx(spec: Spec, out_path: Path, *, pool: Optional[CandidatePool] = None, settings: Optional[Settings] = None) -> Path:
    s = settings or Settings()
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Спецификация"

    headers = ["№", "Производитель", "Артикул", "Описание", "Цена, Р", "Кол-во, шт", "Сумма, ₽", "Комментарий"]
    ws.append(headers)

    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    col_widths = [5, 18, 18, 70, 12, 12, 14, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    by_cat: dict[str, list] = {}
    for it in spec.items:
        by_cat.setdefault(it.category or "misc", []).append(it)

    r = 2
    n = 0
    for cat, items in by_cat.items():
        cat_ru = _CATEGORY_RU.get(cat, cat)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
        cell = ws.cell(r, 1)
        cell.value = cat_ru
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        r += 1

        for it in items:
            n += 1
            price = it.unit_price.amount if it.unit_price else None
            total = (price * it.qty) if price is not None else None
            ev_ids = (it.evidence.bitrix_task_ids or [])[:2]
            ev = " ; ".join(_task_link(tid, pool, s) for tid in ev_ids) if ev_ids else ""
            comment = ev

            ws.cell(r, 1).value = n
            ws.cell(r, 2).value = it.manufacturer or ""
            ws.cell(r, 3).value = it.sku or ""
            ws.cell(r, 4).value = it.description or it.name
            ws.cell(r, 5).value = float(price) if price is not None else None
            ws.cell(r, 6).value = float(it.qty)
            ws.cell(r, 7).value = float(total) if total is not None else None
            ws.cell(r, 8).value = comment

            ws.cell(r, 4).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(r, 8).alignment = Alignment(wrap_text=True, vertical="top")
            r += 1

    wb.save(out_path)
    return out_path
