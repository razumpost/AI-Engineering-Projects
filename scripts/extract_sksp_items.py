from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

import openpyxl


HEADER_SYNONYMS = {
    "name": ["наименование", "название", "позиция", "товар", "номенклат", "name", "item"],
    "desc": ["описание", "характерист", "комментар", "description", "desc"],
    "sku": ["артикул", "код", "sku", "part", "pn", "partnumber", "арт."],
    "qty": ["кол-во", "количество", "qty", "колич", "кол."],
    "unit": ["ед", "ед.", "единиц", "unit", "шт", "шт."],
    "price": ["цена", "стоимость", "price"],
    "sum": ["сумма", "итого", "amount", "total"],
    "vendor": ["производитель", "бренд", "vendor", "brand", "manufacturer"],
}

STOP_WORDS = ["итого", "всего", "total", "sum", "сумма", "subtotal"]


def _norm(v: Any) -> str:
    return str(v or "").strip()


def _as_decimal(val: Any) -> Decimal | None:
    if val is None:
        return None
    if isinstance(val, Decimal):
        return val
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    s = _norm(val).replace("\u00A0", " ").replace(" ", "").replace(",", ".")
    if not s:
        return None
    try:
        return Decimal(s)
    except Exception:
        return None


def _score_header_row(cells: list[str]) -> int:
    score = 0
    row_text = " | ".join(cells).casefold()
    for syns in HEADER_SYNONYMS.values():
        for syn in syns:
            if syn in row_text:
                score += 2
                break
    return score


def _iter_rows_text(sheet, max_rows: int, max_cols: int) -> Iterable[tuple[int, list[str]]]:
    """Yield (row_index_1based, [cell_text...]) for first max_rows."""
    r = 0
    for r_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=1):
        r = r_idx
        cells = [_norm(x) for x in (row or ())[:max_cols]]
        yield r_idx, cells


def _find_header(sheet, max_rows: int = 80, max_cols: int = 60) -> tuple[int | None, dict[str, int]]:
    best_row: int | None = None
    best_score = 0
    best_cells: list[str] = []

    for r_idx, cells in _iter_rows_text(sheet, max_rows=max_rows, max_cols=max_cols):
        score = _score_header_row(cells)
        if score > best_score:
            best_score = score
            best_row = r_idx
            best_cells = cells

    if best_row is None or best_score < 4:
        return None, {}

    col_map: dict[str, int] = {}
    for col_idx, cell in enumerate(best_cells, start=1):
        t = cell.casefold()
        for field, syns in HEADER_SYNONYMS.items():
            if field in col_map:
                continue
            if any(syn in t for syn in syns):
                col_map[field] = col_idx
                break

    if "name" not in col_map:
        return None, {}
    if "qty" not in col_map and "price" not in col_map and "sum" not in col_map:
        return None, {}

    return best_row, col_map


def _row_cell(row_tuple: tuple[Any, ...] | None, col_1based: int | None) -> Any:
    if not row_tuple or not col_1based or col_1based <= 0:
        return None
    idx = col_1based - 1
    if idx >= len(row_tuple):
        return None
    return row_tuple[idx]


def _row_is_empty(row_tuple: tuple[Any, ...] | None, cols_watch: list[int]) -> bool:
    if not row_tuple:
        return True
    for c in cols_watch:
        v = _row_cell(row_tuple, c)
        if _norm(v):
            return False
    return True


def _looks_like_stop_row(name: str) -> bool:
    t = (name or "").casefold()
    return any(w in t for w in STOP_WORDS)


def extract_items_from_xlsx(xlsx_path: Path, max_sheets: int = 12) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    out: list[dict[str, Any]] = []

    for sheet_name in wb.sheetnames[:max_sheets]:
        sh = wb[sheet_name]
        header_row, col_map = _find_header(sh)
        if header_row is None:
            continue

        cols_watch = sorted(set(col_map.values()))
        empty_streak = 0

        # iterate data rows after header
        for r_idx, row in enumerate(
            sh.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            if _row_is_empty(row, cols_watch):
                empty_streak += 1
                if empty_streak >= 25:
                    break
                continue
            empty_streak = 0

            name = _norm(_row_cell(row, col_map.get("name")))
            if not name or _looks_like_stop_row(name):
                continue

            rec = {
                "source_file": str(xlsx_path),
                "sheet": sheet_name,
                "row": r_idx,
                "name": name,
                "description": _norm(_row_cell(row, col_map.get("desc"))) if col_map.get("desc") else None,
                "sku": _norm(_row_cell(row, col_map.get("sku"))) if col_map.get("sku") else None,
                "vendor": _norm(_row_cell(row, col_map.get("vendor"))) if col_map.get("vendor") else None,
                "qty": _as_decimal(_row_cell(row, col_map.get("qty"))) if col_map.get("qty") else None,
                "unit": _norm(_row_cell(row, col_map.get("unit"))) if col_map.get("unit") else None,
                "price": _as_decimal(_row_cell(row, col_map.get("price"))) if col_map.get("price") else None,
                "sum": _as_decimal(_row_cell(row, col_map.get("sum"))) if col_map.get("sum") else None,
            }
            out.append(rec)

    return out


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as f:
        for ln in f:
            ln = ln.strip()
            if not ln:
                continue
            rows.append(json.loads(ln))
    return rows


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--pairs", default="", help="JSONL from build_pairs.py. If set, extracts for each row.")
    ap.add_argument("--xlsx", default="", help="Extract from one xlsx file.")
    ap.add_argument("--out", default="/tmp/sksp_items.jsonl")
    ap.add_argument("--limit-pairs", type=int, default=0)
    ap.add_argument("--skip-errors", action="store_true", help="Skip broken files instead of failing.")
    args = ap.parse_args()

    out_path = Path(args.out).expanduser().resolve()
    out_f = out_path.open("w", encoding="utf-8")

    try:
        if args.xlsx.strip():
            x = Path(args.xlsx).expanduser().resolve()
            items = extract_items_from_xlsx(x)
            for it in items:
                out_f.write(json.dumps(it, ensure_ascii=False, default=str) + "\n")
            print(f"items={len(items)} -> {out_path}")
            return 0

        if args.pairs.strip():
            pairs = _read_jsonl(Path(args.pairs).expanduser().resolve())
            if args.limit_pairs and args.limit_pairs > 0:
                pairs = pairs[: int(args.limit_pairs)]

            total = 0
            skipped = 0

            for p in pairs:
                abs_path = p.get("sksp_abs_path")
                if not abs_path:
                    continue
                x = Path(abs_path)
                suf = x.suffix.casefold()
                if suf not in {".xlsx", ".xlsm"}:
                    continue
                if not x.exists():
                    continue

                try:
                    items = extract_items_from_xlsx(x)
                except Exception as e:
                    skipped += 1
                    msg = f"[skip] {x}: {type(e).__name__}: {e}"
                    if args.skip_errors:
                        print(msg, file=sys.stderr)
                        continue
                    raise

                for it in items:
                    it["deal_id"] = p.get("deal_id")
                    it["sksp_file_id"] = p.get("sksp_file_id")
                    out_f.write(json.dumps(it, ensure_ascii=False, default=str) + "\n")

                total += len(items)

            print(f"total_items={total} skipped_files={skipped} -> {out_path}")
            return 0

        raise SystemExit("Provide --pairs or --xlsx")
    finally:
        out_f.close()


if __name__ == "__main__":
    raise SystemExit(main())